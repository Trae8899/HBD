"""Excel report generation."""

from __future__ import annotations

import json
import zipfile
from pathlib import Path
from typing import Any, Mapping, Sequence

try:  # pragma: no cover - optional dependency
    from openpyxl import Workbook  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore[attr-defined]
except ImportError:  # pragma: no cover - fallback path
    Workbook = None  # type: ignore
    get_column_letter = None  # type: ignore

SUMMARY_ORDER: tuple[tuple[str, tuple[str, ...]], ...] = (
    ("GT Power (MW)", ("summary", "GT_power_MW")),
    ("ST Power (MW)", ("summary", "ST_power_MW")),
    ("Aux Load (MW)", ("summary", "AUX_load_MW")),
    ("Net Power (MW)", ("summary", "NET_power_MW")),
    ("Net Eff. (% LHV)", ("summary", "NET_eff_LHV_pct")),
    ("Stack Temp (°C)", ("hrsg_block", "stack_temp_C")),
    ("Condenser Vacuum (kPa abs)", ("st_block", "lp_section", "exhaust_kPa_abs")),
)


def _resolve_path(root: Mapping[str, Any], path: tuple[str, ...]) -> Any:
    value: Any = root
    for key in path:
        if isinstance(value, Mapping):
            value = value.get(key)
        else:
            return None
    return value


def _write_summary_sheet_openpyxl(workbook: Workbook, result: Mapping[str, Any]) -> None:  # type: ignore[valid-type]
    sheet = workbook.active
    sheet.title = "Summary"
    for row_index, (label, path) in enumerate(SUMMARY_ORDER, start=1):
        sheet.cell(row=row_index, column=1, value=label)
        sheet.cell(row=row_index, column=2, value=_resolve_path(result, path))

    meta = result.get("meta", {})
    start_row = len(SUMMARY_ORDER) + 2
    sheet.cell(row=start_row, column=1, value="Input Case")
    sheet.cell(row=start_row, column=2, value=meta.get("input_case"))
    sheet.cell(row=start_row + 1, column=1, value="Timestamp (UTC)")
    sheet.cell(row=start_row + 1, column=2, value=meta.get("timestamp_utc"))
    sheet.cell(row=start_row + 2, column=1, value="Solver Commit")
    sheet.cell(row=start_row + 2, column=2, value=meta.get("solver_commit"))
    if meta.get("warnings"):
        sheet.cell(row=start_row + 4, column=1, value="Warnings")
        sheet.cell(row=start_row + 4, column=2, value="; ".join(meta["warnings"]))

    sheet.column_dimensions[get_column_letter(1)].width = 32
    sheet.column_dimensions[get_column_letter(2)].width = 22


def _write_stream_sheet_openpyxl(workbook: Workbook, result: Mapping[str, Any]) -> None:  # type: ignore[valid-type]
    sheet = workbook.create_sheet("Streams")
    sheet.append([
        "Block",
        "Descriptor",
        "Pressure (bar abs / kPa abs)",
        "Temperature (°C)",
        "Flow (kg/s)",
        "Power (MW)",
    ])

    hrsg = result.get("hrsg_block", {})
    for level in ("hp", "ip", "lp"):
        level_data = hrsg.get(level, {})
        sheet.append(
            [
                "HRSG",
                level.upper(),
                level_data.get("steam_P_bar"),
                level_data.get("steam_T_C"),
                level_data.get("steam_flow_kg_s"),
                "-",
            ]
        )

    st = result.get("st_block", {})
    for section_key in ("hp_section", "ip_section", "lp_section"):
        section = st.get(section_key, {})
        pressure = section.get("inlet_P_bar")
        if pressure is None:
            pressure = section.get("exhaust_kPa_abs")
        sheet.append(
            [
                "Steam Turbine",
                section_key.replace("_", " ").title(),
                pressure,
                section.get("inlet_T_C"),
                section.get("flow_kg_s"),
                section.get("power_MW"),
            ]
        )

    condenser = result.get("condenser_block", {})
    sheet.append([
        "Condenser",
        "Cooling Water",
        "-",
        condenser.get("cw_outlet_C"),
        condenser.get("lp_flow_kg_s"),
        condenser.get("heat_rejected_MW"),
    ])

    for column in range(1, 7):
        sheet.column_dimensions[get_column_letter(column)].width = 20


def _write_calculation_sheet_openpyxl(workbook: Workbook, trace: Mapping[str, Any]) -> None:  # type: ignore[valid-type]
    sheet = workbook.create_sheet("Calculations")
    sheet.append(["Block", "Detail"])
    for block_name, block_trace in trace.items():
        serialized = json.dumps(block_trace, ensure_ascii=False, indent=2)
        sheet.append([block_name, serialized])
        sheet.row_dimensions[sheet.max_row].height = max(20, min(120, 12 * serialized.count("\n")))
    sheet.column_dimensions[get_column_letter(1)].width = 18
    sheet.column_dimensions[get_column_letter(2)].width = 80


def _export_with_openpyxl(result: Mapping[str, Any], trace: Mapping[str, Any], output_path: Path) -> Path:
    workbook = Workbook()  # type: ignore[valid-type]
    _write_summary_sheet_openpyxl(workbook, result)
    _write_stream_sheet_openpyxl(workbook, result)
    _write_calculation_sheet_openpyxl(workbook, trace)
    workbook.save(output_path)
    return output_path


def _inline_string_cell(cell_ref: str, value: Any) -> str:
    from xml.sax.saxutils import escape

    if value is None:
        text = ""
    else:
        text = escape(str(value))
    return f'<c r="{cell_ref}" t="inlineStr"><is><t>{text}</t></is></c>'


def _sheet_xml(rows: Sequence[Sequence[Any]]) -> str:
    def column_letter(index: int) -> str:
        result = ""
        current = index
        while current > 0:
            current, remainder = divmod(current - 1, 26)
            result = chr(65 + remainder) + result
        return result

    parts = [
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">',
        "<sheetData>",
    ]
    for row_index, row in enumerate(rows, start=1):
        parts.append(f'<row r="{row_index}">')
        for col_index, value in enumerate(row, start=1):
            col_letter = column_letter(col_index)
            parts.append(_inline_string_cell(f"{col_letter}{row_index}", value))
        parts.append("</row>")
    parts.append("</sheetData></worksheet>")
    return "".join(parts)


def _build_fallback_workbook(sheets: Sequence[tuple[str, Sequence[Sequence[Any]]]]) -> dict[str, str]:
    workbook_xml = [
        '<workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" ',
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">',
        "<sheets>",
    ]
    rels = ["<?xml version='1.0' encoding='UTF-8'?><Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'>"]
    content_types = [
        "<?xml version='1.0' encoding='UTF-8'?>",
        "<Types xmlns='http://schemas.openxmlformats.org/package/2006/content-types'>",
        "<Default Extension='rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/>",
        "<Default Extension='xml' ContentType='application/xml'/>",
    ]

    sheet_entries = []
    for index, (title, rows) in enumerate(sheets, start=1):
        sheet_entries.append((f"xl/worksheets/sheet{index}.xml", _sheet_xml(rows)))
        workbook_xml.append(
            f"<sheet name='{title}' sheetId='{index}' r:id='rId{index}'/>"
        )
        rels.append(
            f"<Relationship Id='rId{index}' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet' Target='worksheets/sheet{index}.xml'/>"
        )
        content_types.append(
            f"<Override PartName='/xl/worksheets/sheet{index}.xml' ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'/>"
        )

    workbook_xml.append("</sheets></workbook>")
    rels.append("</Relationships>")
    content_types.append(
        "<Override PartName='/xl/workbook.xml' ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml'/>")
    content_types.append("</Types>")

    data = {
        "[Content_Types].xml": "".join(content_types),
        "_rels/.rels": "<?xml version='1.0' encoding='UTF-8'?><Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'><Relationship Id='rId1' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument' Target='xl/workbook.xml'/></Relationships>",
        "xl/workbook.xml": "".join(workbook_xml),
        "xl/_rels/workbook.xml.rels": "".join(rels),
    }

    for sheet_path, sheet_xml in sheet_entries:
        data[sheet_path] = sheet_xml

    return data


def _export_with_fallback(result: Mapping[str, Any], trace: Mapping[str, Any], output_path: Path) -> Path:
    summary_rows = [[label, _resolve_path(result, path)] for label, path in SUMMARY_ORDER]
    meta = result.get("meta", {})
    summary_rows.extend(
        [
            ["Input Case", meta.get("input_case")],
            ["Timestamp (UTC)", meta.get("timestamp_utc")],
            ["Solver Commit", meta.get("solver_commit")],
        ]
    )
    if meta.get("warnings"):
        summary_rows.append(["Warnings", "; ".join(meta["warnings"])])

    hrsg = result.get("hrsg_block", {})
    st = result.get("st_block", {})
    condenser = result.get("condenser_block", {})

    stream_rows = [
        ["Block", "Descriptor", "Pressure", "Temperature", "Flow", "Power"],
    ]
    for level in ("hp", "ip", "lp"):
        level_data = hrsg.get(level, {})
        stream_rows.append(
            [
                "HRSG",
                level.upper(),
                level_data.get("steam_P_bar"),
                level_data.get("steam_T_C"),
                level_data.get("steam_flow_kg_s"),
                "-",
            ]
        )
    for section_key in ("hp_section", "ip_section", "lp_section"):
        section = st.get(section_key, {})
        pressure = section.get("inlet_P_bar")
        if pressure is None:
            pressure = section.get("exhaust_kPa_abs")
        stream_rows.append(
            [
                "Steam Turbine",
                section_key.replace("_", " ").title(),
                pressure,
                section.get("inlet_T_C"),
                section.get("flow_kg_s"),
                section.get("power_MW"),
            ]
        )
    stream_rows.append(
        [
            "Condenser",
            "Cooling Water",
            "-",
            condenser.get("cw_outlet_C"),
            condenser.get("lp_flow_kg_s"),
            condenser.get("heat_rejected_MW"),
        ]
    )

    calc_rows = [["Block", "Detail"]]
    for block_name, block_trace in trace.items():
        calc_rows.append([block_name, json.dumps(block_trace, ensure_ascii=False, indent=2)])

    workbook_parts = _build_fallback_workbook(
        (
            ("Summary", summary_rows),
            ("Streams", stream_rows),
            ("Calculations", calc_rows),
        )
    )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(output_path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        for internal_path, xml_content in workbook_parts.items():
            archive.writestr(internal_path, xml_content)
    return output_path


def export_summary_to_excel(result: Mapping[str, Any], trace: Mapping[str, Any], output_path: Path) -> Path:
    """Create an Excel workbook representing the HBD summary and stream tables."""
    if Workbook is not None:  # type: ignore[truthy-function]
        output_path.parent.mkdir(parents=True, exist_ok=True)
        return _export_with_openpyxl(result, trace, output_path)
    return _export_with_fallback(result, trace, output_path)
